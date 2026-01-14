"""Excel处理核心逻辑 - 动态识别表头行数，首文件保留表头，后续文件剔除表头"""
import os
import re
import xlrd
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
from .config import Config
from .utils import Utils
from .format_handler import FormatHandler

class ExcelProcessor:
    """Excel/CSV文件处理核心类 - 动态识别表头行数（无时间=表头，有时间=数据开始）"""
    def __init__(self, folder_path, output_path, log_callback):
        self.folder_path = folder_path
        self.output_path = output_path
        self.log = log_callback  # 日志回调函数
        self.excel_files = []
        self.converted_files = []  # 存储转换后的文件路径
        self.header_info = []  # (列索引, 原始表头, 标准化表头, 是否金额列)
        self.header_map = {}
        self.wb = None
        self.ws = None
        self.has_shop_column = False  # 是否已添加店铺列
        self.header_rows = None  # 动态识别，初始化为None
        self.header_end_row = None  # 表头结束行号（动态赋值）
        self.base_first_column_orig = ""
        self.base_first_column_norm = ""
        self.clean_base_header = []
        self.keyword_match_threshold = 0.9  # 表头判定阈值

        # 初始化时过滤警告
        Utils.filter_warnings()

    def merge(self):
        """执行合并操作 - 先动态识别表头，再处理文件"""
        self._get_excel_files_in_native_order()
        self._convert_all_xls_to_xlsx()
        first_file = self.converted_files[0]

        # 核心步骤1：动态识别第一个文件的表头行数
        self._detect_header_rows_auto(first_file)
        self.log(f"动态识别完成，表头行数(header_rows)={self.header_rows}")

        first_row_count, start_row, format_ref_row = self._analyze_first_file(first_file)
        self._add_shop_column_to_first_file(first_file, first_row_count)
        merged_df = self._process_all_files_in_native_order()
        self._write_merged_data(merged_df, first_row_count, start_row, format_ref_row)
        result = self._save_result()
        self._cleanup_converted_files()
        return result

    def _detect_header_rows_auto(self, first_file):
        """
        核心方法：动态识别表头行数
        逻辑：从第1行开始扫描，无时间数据=表头，有时间数据=数据开始，确定header_rows
        """
        self.log(f"开始动态扫描第一个文件 '{os.path.basename(first_file)}'，识别表头行数...")
        row_is_data = False
        header_candidate = 0  # 表头行数候选值

        try:
            # 读取第一个文件的所有行（无表头模式）
            if Utils.is_csv_file(first_file):
                df_all = pd.read_csv(first_file, header=None, dtype=str, keep_default_na=False)
            else:
                # 先转换xls（如果需要），再读取
                if first_file.endswith('.xls') and not first_file.endswith('.xlsx'):
                    first_file = self._convert_xls_to_xlsx(first_file)
                df_all = pd.read_excel(first_file, header=None, dtype=str, keep_default_na=False)

            if df_all.empty:
                raise IOError("第一个文件无任何数据，无法动态识别表头")

            # 逐行扫描（从第1行开始，对应索引0）
            for row_idx, (_, row) in enumerate(df_all.iterrows()):
                row_data = row.values
                row_str = " ".join([str(val).strip() for val in row_data if str(val).strip() != ""])

                # 核心判断：该行是否包含有效时间/日期数据
                if self._has_valid_time_data(row_str):
                    row_is_data = True
                    header_candidate = row_idx  # 数据行的前所有行都是表头（行号=索引）
                    self.log(f"  第{row_idx+1}行检测到有效时间数据，判定为数据开始行")
                    break
                else:
                    self.log(f"  第{row_idx+1}行无时间数据，判定为表头行")

            # 确定最终header_rows
            if row_is_data:
                self.header_rows = header_candidate  # 数据行索引=表头行数（前header_candidate行是表头）
            else:
                # 边界兜底：无任何时间数据，默认设为2行表头（兼容原有场景）
                self.log("  警告：文件中未检测到任何时间数据，默认表头行数=2")
                self.header_rows = 2

            # 同步更新表头结束行号
            self.header_end_row = self.header_rows

        except Exception as e:
            self.log(f"动态识别表头行数失败：{str(e)}，默认表头行数=2")
            self.header_rows = 2
            self.header_end_row = 2

    def _has_valid_time_data(self, content):
        """
        辅助方法：判断字符串是否包含有效时间/日期数据
        兼容格式：yyyy-mm-dd、yyyy/mm/dd、yyyy.mm.dd、hh:mm:ss、yyyy年mm月dd日
        """
        if not content or len(content) < 6:
            return False

        # 正则匹配常见时间/日期格式（优先级从高到低）
        time_patterns = [
            r'\d{4}[-/\.]\d{1,2}[-/\.]\d{1,2}',  # yyyy-mm-dd / yyyy/mm/dd / yyyy.mm.dd
            r'\d{1,2}[-/\.]\d{1,2}[-/\.]\d{4}',  # dd-mm-yyyy / mm/dd/yyyy
            r'\d{2}:\d{2}:\d{2}',  # hh:mm:ss
            r'\d{4}年\d{1,2}月\d{1,2}日'  # 中文日期：yyyy年mm月dd日
        ]

        for pattern in time_patterns:
            if re.search(pattern, content):
                return True

        # 额外判断：是否包含datetime可转换的字符串（兜底）
        try:
            # 尝试转换常见格式，避免正则漏判
            datetime.strptime(content, "%Y-%m-%d")
            return True
        except:
            pass
        try:
            datetime.strptime(content, "%Y/%m/%d")
            return True
        except:
            pass

        return False

    def _convert_all_xls_to_xlsx(self):
        """将所有xls文件转换为xlsx格式，CSV文件直接跳过"""
        self.log("开始检查并转换xls文件...")
        self.converted_files = []
        for file in self.excel_files:
            file_path = os.path.join(self.folder_path, file)
            if Utils.is_csv_file(file_path):
                self.converted_files.append(file_path)
                continue
            if file.endswith('.xls') and not file.endswith('.xlsx'):
                converted_path = self._convert_xls_to_xlsx(file_path)
                self.converted_files.append(converted_path)
            else:
                self.converted_files.append(file_path)
        self.log(f"格式转换完成，共处理{len(self.converted_files)}个文件")

    def _convert_xls_to_xlsx(self, xls_file_path):
        """将单个xls文件转换为xlsx格式"""
        try:
            file_name = os.path.basename(xls_file_path)
            file_name_without_ext = os.path.splitext(file_name)[0]
            xlsx_file_path = os.path.join(self.folder_path, f"{file_name_without_ext}_temp_converted.xlsx")
            workbook = xlrd.open_workbook(xls_file_path)
            sheet = workbook.sheet_by_index(0)
            new_workbook = Workbook()
            new_sheet = new_workbook.active
            for row_idx in range(sheet.nrows):
                row_data = sheet.row_values(row_idx)
                new_sheet.append(row_data)
            new_workbook.save(xlsx_file_path)
            self.log(f"已将 '{file_name}' 从xls转换为xlsx格式")
            return xlsx_file_path
        except Exception as e:
            self.log(f"转换文件 '{xls_file_path}' 失败: {str(e)}")
            raise

    def _cleanup_converted_files(self):
        """清理临时转换的xlsx文件，跳过CSV文件"""
        try:
            for file_path in self.converted_files:
                if Utils.is_csv_file(file_path):
                    continue
                if "_temp_converted.xlsx" in file_path:
                    os.remove(file_path)
                    self.log(f"已清理临时文件: {os.path.basename(file_path)}")
        except Exception as e:
            self.log(f"清理临时文件时出错: {str(e)}")

    def _get_excel_files_in_native_order(self):
        """获取文件夹中所有Excel/CSV文件，保持操作系统原生顺序"""
        self.log(f"正在扫描文件夹: {self.folder_path}")
        self.excel_files = [
            f for f in os.listdir(self.folder_path)
            if (f.endswith(Config.EXCEL_EXTENSIONS) or Utils.is_csv_file(f))
            and not f.startswith(Config.TEMP_FILE_PREFIX)
        ]
        if not self.excel_files:
            raise FileNotFoundError("未找到任何Excel/CSV文件")
        self.log(f"找到{len(self.excel_files)}个Excel/CSV文件，将按以下原生顺序处理:")
        for i, file in enumerate(self.excel_files, 1):
            self.log(f"  {i}. {file}")

    def _add_shop_column_to_first_file(self, first_file, row_count):
        """为动态识别的表头行数添加店铺列，保留所有表头行"""
        if not self.ws or self.has_shop_column or self.header_rows is None:
            return
        self.ws.insert_cols(1)
        # 给所有表头行都填充"店铺"列名（动态适配header_rows）
        for header_row in range(1, self.header_rows + 1):
            shop_header_cell = self.ws.cell(row=header_row, column=1)
            shop_header_cell.value = "店铺"
            if self.ws.max_column >= 2:
                ref_cell = self.ws.cell(row=header_row, column=2)
                FormatHandler.copy_cell_format(ref_cell, shop_header_cell)
        self.log(f"已为第一个文件的{self.header_rows}行表头添加店铺列")

        # 填充第一个文件的数据行店铺值
        shop_name = os.path.splitext(os.path.basename(first_file))[0].replace("_temp_converted", "")
        data_start_row = self.header_end_row + 1  # 数据从表头结束行+1开始
        for row in range(data_start_row, data_start_row + row_count):
            if row > self.ws.max_row:
                break
            cell = self.ws.cell(row=row, column=1)
            cell.value = shop_name
            if self.ws.max_column >= 2:
                ref_cell = self.ws.cell(row=row, column=2)
                FormatHandler.copy_cell_format(ref_cell, cell)
        self.header_info.insert(0, (1, "店铺", "店铺", False))
        self.header_map["店铺"] = 1
        for i in range(1, len(self.header_info)):
            col_idx, orig_header, normalized, is_amount_col = self.header_info[i]
            self.header_info[i] = (col_idx + 1, orig_header, normalized, is_amount_col)
            if normalized in self.header_map:
                self.header_map[normalized] = col_idx + 1
        self.has_shop_column = True

    def _analyze_first_file(self, first_file):
        """基于动态识别的header_rows，分析第一个文件"""
        if self.header_rows is None:
            raise ValueError("表头行数未动态识别，无法分析文件")

        self.log(f"以第一个文件 '{os.path.basename(first_file)}' 为基础，分析{self.header_rows}行表头")

        if Utils.is_csv_file(first_file):
            df_all = pd.read_csv(first_file, header=None, dtype=str, keep_default_na=False)
            if df_all.empty:
                raise IOError(f"第一个文件{os.path.basename(first_file)}是CSV格式但无数据")
            # 提取最后一行表头作为有效表头（索引=header_rows-1）
            header_row_idx = self.header_rows - 1
            effective_header = df_all.iloc[header_row_idx].tolist()
            # 保留所有表头行，数据从header_rows开始
            df_header = df_all.iloc[:self.header_rows].reset_index(drop=True)
            df_data = df_all.iloc[self.header_rows:].reset_index(drop=True)
            df_data.columns = effective_header
            temp_xlsx = os.path.join(self.folder_path, "_temp_first_file.xlsx")
            with pd.ExcelWriter(temp_xlsx, engine="openpyxl") as writer:
                df_header.to_excel(writer, sheet_name="Sheet", index=False, header=False)
                df_data.to_excel(writer, sheet_name="Sheet", index=False, header=False, startrow=self.header_rows)
            try:
                self.wb = load_workbook(temp_xlsx)
            except Exception as e:
                raise IOError(f"加载CSV转换的临时文件失败: {str(e)}")
            finally:
                if os.path.exists(temp_xlsx):
                    os.remove(temp_xlsx)
        else:
            try:
                self.wb = load_workbook(first_file)
            except Exception as e:
                raise IOError(f"无法加载文件 {os.path.basename(first_file)}: {str(e)}")

        self.ws = self.wb.active
        self.log(f"第一个文件表头行数：{self.header_rows}，表头结束行号：{self.header_end_row}")

        col_idx = 1
        self.header_info = []
        self.header_map = {}
        if Utils.is_csv_file(first_file):
            for orig_header in effective_header:
                if col_idx > Config.MAX_COLUMNS_TO_CHECK:
                    break
                normalized = Utils.normalize_header(orig_header)
                is_amount_col = Utils.is_amount_column(str(orig_header)) if orig_header else False
                self.header_info.append((col_idx, orig_header, normalized, is_amount_col))
                if normalized and normalized not in self.header_map:
                    self.header_map[normalized] = col_idx
                col_idx += 1
        else:
            while True:
                original_header = self.ws.cell(row=self.header_end_row, column=col_idx).value
                normalized = Utils.normalize_header(original_header)
                is_amount_col = Utils.is_amount_column(str(original_header)) if original_header else False
                self.header_info.append((col_idx, original_header, normalized, is_amount_col))
                if normalized and normalized not in self.header_map:
                    self.header_map[normalized] = col_idx
                if self._is_header_end(col_idx) or col_idx >= Config.MAX_COLUMNS_TO_CHECK:
                    break
                col_idx += 1

        # 生成基准表头关键词（用于后续文件过滤）
        self.clean_base_header = [
            str(h[1]).strip().lower() for h in self.header_info
            if h[1] is not None and str(h[1]).strip() != ""
        ]
        self.log(f"已生成基准表头关键词(第一个文件第{self.header_end_row}行表头): {self.clean_base_header[:5]}...")

        # 记录基准非店铺第一列
        if len(self.header_info) > 0:
            self.base_first_column_orig = self.header_info[0][1]
            self.base_first_column_norm = self.header_info[0][2]
            self.log(f"已记录基准第一列：原始='{self.base_first_column_orig}'，标准化='{self.base_first_column_norm}'")

        self.log(f"第一个文件检测到的表头信息（共 {len(self.header_info)} 列）:")
        for idx, orig, norm, _ in self.header_info:
            self.log(f"  第{idx}列: 原始='{orig}'，标准化='{norm}'")

        try:
            if Utils.is_csv_file(first_file):
                df_all = pd.read_csv(first_file, header=None, dtype=str)
                first_df = df_all.iloc[self.header_rows:].reset_index(drop=True)
                first_df.columns = effective_header
            else:
                first_df = pd.read_excel(first_file, skiprows=self.header_rows, header=None, dtype=str)
                first_df.columns = [h[1] for h in self.header_info[:len(first_df.columns)]]
        except Exception as e:
            raise IOError(f"无法读取第一个文件数据: {str(e)}")

        first_row_count = len(first_df)
        start_row = self.header_end_row + first_row_count + 1
        format_ref_row = self.header_end_row + 1 if first_row_count > 0 else self.header_end_row
        return first_row_count, start_row, format_ref_row

    def _is_header_end(self, col_idx):
        """判断表头是否结束（列维度）"""
        empty_count = 0
        for i in range(1, Config.EMPTY_COLUMN_THRESHOLD + 1):
            if col_idx + i > Config.MAX_COLUMNS_TO_CHECK:
                return True
            if self.ws.cell(row=self.header_end_row, column=col_idx + i).value is None:
                empty_count += 1
        return empty_count >= Config.EMPTY_COLUMN_THRESHOLD

    def _is_header_row(self, row_data):
        """严格判定：只有行内容和表头高度一致才判定为表头行"""
        if len(self.clean_base_header) == 0 or len(row_data) < 3:
            return False

        row_str = " ".join([str(val).strip().lower() for val in row_data if pd.notna(val) and str(val).strip() != ""])
        if not row_str:
            return False

        match_count = sum([1 for keyword in self.clean_base_header if keyword in row_str])
        match_ratio = match_count / len(self.clean_base_header) if len(self.clean_base_header) > 0 else 0

        has_digit = any(char.isdigit() for char in row_str)
        is_plain_text = not has_digit

        # 只有 匹配度≥90% 且 纯文本 才判定为表头行
        return match_ratio >= self.keyword_match_threshold and is_plain_text

    def _filter_header_rows(self, df, file_name, file_index):
        """后续文件过滤表头行，保留数据不丢失"""
        if df.empty:
            return df

        # 第一个文件直接返回，不做任何过滤
        if file_index == 0:
            self.log(f"  第一个文件 {file_name} 保留所有数据行")
            return df

        original_len = len(df)
        if original_len == 0:
            return df

        header_mask = [True] * original_len
        header_row_count = 0

        # 只剔除前2行中明显的表头行，避免删光数据
        for idx in range(min(2, original_len)):
            row = df.iloc[idx]
            if self._is_header_row(row.values):
                header_mask[idx] = False
                header_row_count += 1
                self.log(f"  [精准剔除] {file_name} 表头行[{idx+1}]: {row.values[:3]}...")

        filtered_df = df[header_mask].reset_index(drop=True)
        filtered_len = len(filtered_df)

        # 兜底：过滤后无数据则恢复原始数据
        if filtered_len == 0:
            self.log(f"  警告：{file_name}过滤后无数据，恢复原始{original_len}行数据")
            return df

        self.log(f"  {file_name} 过滤前行数: {original_len} → 过滤后行数: {filtered_len} (剔除{header_row_count}行表头)")
        return filtered_df

    def _read_csv_file_for_merge(self, file_path, file_index):
        """读取CSV文件，基于动态header_rows跳过表头"""
        try:
            file_name = os.path.basename(file_path)
            self.log(f"读取CSV文件: {file_name} (跳过前{self.header_rows}行表头)")
            df_all = pd.read_csv(file_path, header=None, dtype=str, keep_default_na=False)
            total_rows = len(df_all)

            # 放宽判断：只要总行数 > header_rows 就认为有数据
            if total_rows <= self.header_rows:
                self.log(f"  警告：{file_name}总行数{total_rows}≤表头行数{self.header_rows}，尝试读取所有行")
                df_data = df_all.reset_index(drop=True)
            else:
                header_row_idx = self.header_rows - 1
                file_header = df_all.iloc[header_row_idx].tolist()
                df_data = df_all.iloc[self.header_rows:].reset_index(drop=True)
                df_data.columns = file_header[:len(df_data.columns)]

            self.log(f"  {file_name} 原始数据行数: {len(df_data)}")
            # 过滤表头行（只删前2行）
            df_data = self._filter_header_rows(df_data, file_name, file_index)
            if df_data.empty:
                self.log(f"  {file_name} 过滤后无数据，跳过")
                return None

            # 处理店铺列
            shop_name = os.path.splitext(file_name)[0].replace("_temp_converted", "")
            if "店铺" in df_data.columns:
                df_data["店铺"] = shop_name
            else:
                df_data.insert(0, "店铺", shop_name)

            # 打印后续文件数据预览
            if file_index > 0 and len(df_data) > 0:
                self.log(f"  {file_name} 数据预览(前2行): {df_data.head(2).values.tolist()[:2]}")

            return df_data
        except Exception as e:
            self.log(f"读取CSV文件{file_path}失败: {str(e)}")
            return None

    def _process_all_files_in_native_order(self):
        """按顺序处理所有文件，兼容动态header_rows"""
        all_data = []
        base_orig_to_norm = {h[1]: h[2] for h in self.header_info if h[1] is not None}
        base_columns = [h[2] for h in self.header_info]

        for file_idx, file_path in enumerate(self.converted_files):
            file_name = os.path.basename(file_path)
            try:
                if Utils.is_csv_file(file_path):
                    df = self._read_csv_file_for_merge(file_path, file_idx)
                    if df is None or df.empty:
                        self.log(f"  跳过无数据文件: {file_name}")
                        continue
                else:
                    df = pd.read_excel(file_path, skiprows=self.header_rows, header=None, dtype=str)
                    df.columns = [h[1] for h in self.header_info[:len(df.columns)]]
                    df = self._filter_header_rows(df, file_name, file_idx)
                    if df.empty:
                        self.log(f"  跳过无数据文件: {file_name}")
                        continue

                    shop_name = os.path.splitext(file_name)[0].replace("_temp_converted", "")
                    if "店铺" in df.columns:
                        df["店铺"] = shop_name
                    else:
                        df.insert(0, '店铺', shop_name)

                # 列映射兼容，确保数据不丢失
                df_norm_cols = [Utils.normalize_header(col) for col in df.columns]
                aligned_df = pd.DataFrame(columns=base_columns)
                aligned_df["店铺"] = df["店铺"].values

                # 填充其他列：优先匹配，不匹配则填空值
                for base_col in base_columns:
                    if base_col == "店铺":
                        continue
                    match_col_idx = next((i for i, col in enumerate(df_norm_cols) if col == base_col), None)
                    if match_col_idx is not None:
                        aligned_df[base_col] = df.iloc[:, match_col_idx].values
                    else:
                        aligned_df[base_col] = ""
                        self.log(f"  {file_name} 缺少列 '{base_col}'，填充空值")

                all_data.append(aligned_df)
                self.log(f"处理文件 {file_name} 成功，有效数据行数：{len(aligned_df)}")

            except Exception as e:
                self.log(f"警告: 处理{file_name}出错 - {str(e)}")

        if not all_data:
            raise ValueError("没有可处理的有效文件")

        merged_df = pd.concat(all_data, ignore_index=True)
        self.log(f"\n数据合并完成，共 {len(merged_df)} 行数据，{len(merged_df.columns)} 列")
        return merged_df

    def _write_merged_data(self, merged_df, first_row_count, start_row, format_ref_row):
        """写入合并数据，保留第一个文件的动态表头"""
        if self.ws.max_row >= start_row:
            try:
                self.ws.delete_rows(start_row, self.ws.max_row - start_row + 1)
                self.log("已清除基础文件后的冗余数据（表头未被修改）")
            except Exception as e:
                self.log(f"警告: 清除旧数据时出错 - {str(e)}")
        total_rows = len(merged_df)
        batch_size = Config.WRITE_BATCH_SIZE
        for batch_start in range(first_row_count, total_rows, batch_size):
            batch_end = min(batch_start + batch_size, total_rows)
            self.log(f"正在写入数据: {batch_end}/{total_rows} 行")
            for row_idx in range(batch_start, batch_end):
                data_row = merged_df.iloc[row_idx]
                current_row = start_row + (row_idx - first_row_count)
                for col_info in self.header_info:
                    self._write_cell(data_row, col_info, current_row, format_ref_row)

    def _write_cell(self, data_row, col_info, current_row, format_ref_row):
        """写入单个单元格数据"""
        col_idx, orig_header, norm_header, is_amount_col = col_info
        try:
            value = data_row[norm_header] if norm_header in data_row else ""
            if pd.isna(value):
                value = ""
        except:
            value = ""
        ref_cell = self.ws.cell(row=format_ref_row, column=col_idx)
        target_cell = self.ws.cell(row=current_row, column=col_idx)
        if is_amount_col:
            value = self._process_amount_value(value, target_cell, ref_cell)
        elif str(value).isdigit() and len(str(value)) > Config.LONG_NUMBER_THRESHOLD:
            target_cell.number_format = '@'
            value = str(value)
        else:
            target_cell.number_format = ref_cell.number_format
        target_cell.value = value
        FormatHandler.copy_cell_format(ref_cell, target_cell, force_right=is_amount_col)

    def _process_amount_value(self, value, target_cell, ref_cell):
        """处理金额列的值"""
        try:
            clean_value = str(value).replace(',', '').replace('￥', '').replace('$', '')
            value = float(clean_value)
            target_cell.number_format = ref_cell.number_format
            return value
        except:
            target_cell.number_format = '@'
            return value

    def _save_result(self):
        """保存合并结果，保留动态识别的表头"""
        Utils.ensure_dir_exists(self.output_path)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(self.output_path, f"汇总结果_{timestamp}.xlsx")
        try:
            self.wb.save(output_file)
            self.log(f"合并成功！结果已保存至:\n{output_file}")
            return output_file
        except Exception as e:
            raise IOError(f"保存文件失败: {str(e)}")