"""Excel文件处理核心逻辑"""

import os
import pandas as pd
from openpyxl import load_workbook
from .config import Config
from .utils import Utils
from .format_handler import FormatHandler

class ExcelProcessor:
    """Excel文件处理和写入工具"""
    def __init__(self, folder_path: str = '.', output_dir: str = Config.DEFAULT_OUTPUT_DIR):
        self.folder_path = self._validate_path(folder_path)
        self.output_dir = self._validate_path(output_dir)
        self.excel_files = []
        self.header_info = []  # (列索引, 原始表头, 标准化表头, 是否金额列)
        self.header_map = {}
        self.wb = None
        self.ws = None

    def _validate_path(self, path: str) -> str:
        """验证路径有效性"""
        resolved_path = os.path.abspath(path)
        if not os.path.exists(resolved_path):
            raise NotADirectoryError(f"路径不存在: {resolved_path}")
        if not os.path.isdir(resolved_path):
            raise NotADirectoryError(f"不是有效的目录: {resolved_path}")
        return resolved_path

    def get_excel_files(self) -> None:
        """获取文件夹中所有Excel文件"""
        self.excel_files = [
            f for f in os.listdir(self.folder_path) 
            if f.endswith(Config.EXCEL_EXTENSIONS) and not f.startswith(Config.TEMP_FILE_PREFIX)
        ]
        
        if not self.excel_files:
            raise FileNotFoundError(f"在 {self.folder_path} 中未找到任何Excel文件")
            
        print(f"找到{len(self.excel_files)}个Excel文件，准备合并...")

    def analyze_first_file(self, first_file: str):
        """分析第一个文件，获取表头信息和格式参考"""
        first_file_path = os.path.join(self.folder_path, first_file)
        
        if not os.path.exists(first_file_path):
            raise FileNotFoundError(f"基础文件不存在: {first_file_path}")
            
        print(f"以文件 '{first_file}' 为基础进行合并")
        
        # 加载第一个文件
        try:
            self.wb = load_workbook(first_file_path)
        except Exception as e:
            raise IOError(f"无法加载文件 {first_file}: {str(e)}")
            
        self.ws = self.wb.active
        
        # 分析表头
        col_idx = 1
        while True:
            original_header = self.ws.cell(row=1, column=col_idx).value
            normalized = Utils.normalize_header(original_header)
            is_amount_col = Utils.is_amount_column(str(original_header), Config.AMOUNT_KEYWORDS) if original_header else False
            
            self.header_info.append((col_idx, original_header, normalized, is_amount_col))
            
            if normalized and normalized not in self.header_map:
                self.header_map[normalized] = col_idx
            
            # 检测表头结束
            if self._is_header_end(col_idx):
                break
                
            col_idx += 1
        
        # 显示检测到的金额列
        self._display_amount_columns()
        
        # 读取第一个文件数据
        try:
            first_df = pd.read_excel(first_file_path, dtype=str)
        except Exception as e:
            raise IOError(f"无法读取文件 {first_file}: {str(e)}")
            
        first_row_count = len(first_df)
        start_row = first_row_count + 2  # 数据开始追加的位置
        format_ref_row = 2 if first_row_count > 0 else 1
        
        return first_row_count, start_row, format_ref_row

    def _is_header_end(self, col_idx: int) -> bool:
        """判断表头是否结束"""
        empty_count = 0
        for i in range(1, Config.EMPTY_COLUMN_THRESHOLD + 1):
            if self.ws.cell(row=1, column=col_idx + i).value is None:
                empty_count += 1
        return empty_count >= Config.EMPTY_COLUMN_THRESHOLD

    def _display_amount_columns(self) -> None:
        """显示检测到的金额列"""
        print(f"基础文件表头分析完成: 共 {len(self.header_info)} 列")
        amount_columns = [
            f"第{idx}列: {orig}" 
            for idx, orig, _, is_amt in self.header_info 
            if is_amt
        ]
        if amount_columns:
            print(f"检测到金额相关列: {', '.join(amount_columns)}")

    def process_other_files(self, first_file: str) -> pd.DataFrame:
        """处理其他文件并合并数据"""
        try:
            all_data = [pd.read_excel(os.path.join(self.folder_path, first_file), dtype=str)]
        except Exception as e:
            raise IOError(f"无法读取基础文件 {first_file}: {str(e)}")
        
        for file in self.excel_files[1:]:
            file_path = os.path.join(self.folder_path, file)
            try:
                df = pd.read_excel(file_path, dtype=str)
                file_headers = [Utils.normalize_header(col) for col in df.columns]
                
                # 对齐列
                aligned_df = pd.DataFrame(columns=[h[2] for h in self.header_info])
                for norm_header in aligned_df.columns:
                    if norm_header in file_headers:
                        df_col_idx = file_headers.index(norm_header)
                        aligned_df[norm_header] = df.iloc[:, df_col_idx]
                    else:
                        aligned_df[norm_header] = ""
                
                all_data.append(aligned_df)
                print(f"已处理文件: {file}")
                
            except Exception as e:
                print(f"警告: 处理文件{file}时出错，已跳过该文件 - {str(e)}")
        
        if len(all_data) <= 1:
            print("警告: 仅找到一个有效文件或所有其他文件处理失败")
        
        # 合并所有数据
        merged_df = pd.concat(all_data, ignore_index=True)
        print(f"数据合并完成，共 {len(merged_df)} 行数据")
        return merged_df

    def write_merged_data(
        self, 
        merged_df: pd.DataFrame, 
        first_row_count: int, 
        start_row: int, 
        format_ref_row: int
    ) -> None:
        """将合并后的数据写入Excel"""
        # 清除原有多余数据
        if self.ws.max_row >= start_row:
            try:
                self.ws.delete_rows(start_row, self.ws.max_row - start_row + 1)
                print("已清除基础文件后的原有数据")
            except Exception as e:
                print(f"警告: 清除旧数据时出错 - {str(e)}")
        
        # 写入数据
        for row_idx in range(first_row_count, len(merged_df)):
            data_row = merged_df.iloc[row_idx]
            current_row = start_row + (row_idx - first_row_count)
            
            for col_info in self.header_info:
                self._write_cell(data_row, col_info, current_row, format_ref_row)

    def _write_cell(self, data_row: pd.Series, col_info, current_row: int, format_ref_row: int) -> None:
        """写入单个单元格数据"""
        col_idx, orig_header, norm_header, is_amount_col = col_info
        
        # 获取单元格值
        value = data_row[norm_header] if norm_header in data_row else ""
        if pd.isna(value):
            value = ""
        
        # 获取参考单元格和目标单元格
        ref_cell = self.ws.cell(row=format_ref_row, column=col_idx)
        target_cell = self.ws.cell(row=current_row, column=col_idx)
        
        # 处理金额列
        if is_amount_col:
            value = self._process_amount_value(value, target_cell, ref_cell)
        # 处理长数字
        elif str(value).isdigit() and len(str(value)) > Config.LONG_NUMBER_THRESHOLD:
            target_cell.number_format = '@'
            value = str(value)
        else:
            target_cell.number_format = ref_cell.number_format
        
        target_cell.value = value
        FormatHandler.copy_cell_format(ref_cell, target_cell, force_right=is_amount_col)

    def _process_amount_value(self, value: str, target_cell, ref_cell) -> any:
        """处理金额列的值"""
        try:
            # 清除货币符号和千位分隔符
            clean_value = str(value).replace(',', '').replace('￥', '').replace('$', '')
            value = float(clean_value)
            target_cell.number_format = ref_cell.number_format
            return value
        except:
            # 转换失败时保持文本格式但强制右对齐
            target_cell.number_format = '@'
            return value

    def save_result(self) -> None:
        """保存合并结果"""
        # 确保输出目录存在
        Utils.ensure_dir_exists(self.output_dir)
        
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(self.output_dir, f"合并结果_{timestamp}.xlsx")
        
        try:
            self.wb.save(output_file)
            print(f"\n合并成功！结果文件已保存至:")
            print(f"{output_file}")
        except Exception as e:
            raise IOError(f"保存文件失败: {str(e)}")
